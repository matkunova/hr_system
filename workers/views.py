from rest_framework import viewsets
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status
import logging
from io import BytesIO
from openpyxl import load_workbook
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.db import transaction

from .models import Worker
from .serializers import WorkerListSerializer, WorkerDetailSerializer
from .permissions import WorkerPermission

logger = logging.getLogger(__name__)


class WorkerViewSet(viewsets.ModelViewSet):
    queryset = Worker.objects.all()
    permission_classes = [WorkerPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active', 'position']
    search_fields = ['first_name', 'last_name', 'email']

    def get_serializer_class(self):
        if self.action == 'list':
            return WorkerListSerializer
        return WorkerDetailSerializer

    def perform_create(self, serializer):
        worker = serializer.save(created_by=self.request.user)
        logger.info(f"Worker created via API: {worker.email} by {self.request.user.username}")


class WorkerImportView(APIView):
    """
    Импорт работников из Excel-файла (.xlsx)
    Ожидает файл с полями: first_name, last_name, email, position
    middle_name — опционально
    """
    permission_classes = [IsAdminUser]

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            return Response(
                {"error": "Файл не предоставлен"},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file.name.endswith('.xlsx'):
            return Response(
                {"error": "Поддерживается только .xlsx"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Загружаем Excel
            file.seek(0)
            in_memory_file = BytesIO(file.read())
            workbook = load_workbook(filename=in_memory_file, read_only=True)
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                return Response({"error": "Файл пуст"}, status=status.HTTP_400_BAD_REQUEST)

            # Парсим заголовок
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            logger.info(f"Raw headers from file: {rows[0]}")
            logger.info(f"Processed headers: {headers}")
            required_fields = {'first_name', 'last_name', 'email', 'position'}
            missing = required_fields - set(headers)
            if missing:
                return Response(
                    {"error": f"Отсутствуют столбцы: {', '.join(missing)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            col_index = {name: i for i, name in enumerate(headers)}

            # Сбор данных и валидация
            workers_to_create = []
            errors = []
            seen_emails = set()

            for row_num, row in enumerate(rows[1:], start=2):
                if not any(row):  # пропускаем пустые строки
                    continue

                # Извлекаем данные
                def get_val(field):
                    idx = col_index.get(field, -1)
                    if idx == -1 or idx >= len(row):
                        return ''
                    val = row[idx]
                    return str(val).strip() if val is not None else ''

                first_name = get_val('first_name')
                last_name = get_val('last_name')
                email = get_val('email')
                position = get_val('position')
                middle_name = get_val('middle_name')

                # Обязательные поля
                if not first_name:
                    errors.append({"row": row_num, "error": "Отсутствует first_name"})
                    continue
                if not last_name:
                    errors.append({"row": row_num, "error": "Отсутствует last_name"})
                    continue
                if not email:
                    errors.append({"row": row_num, "error": "Отсутствует email"})
                    continue
                if not position:
                    errors.append({"row": row_num, "error": "Отсутствует position"})
                    continue

                # Валидация email
                try:
                    validate_email(email)
                except ValidationError:
                    errors.append({"row": row_num, "error": f"Некорректный email: {email}"})
                    continue

                # Дубликаты в файле
                if email in seen_emails:
                    errors.append({"row": row_num, "error": f"Дубликат email в файле: {email}"})
                    continue
                seen_emails.add(email)

                # Дубликаты в БД
                if Worker.objects.filter(email=email).exists():
                    errors.append({"row": row_num, "error": f"Email уже существует: {email}"})
                    continue

                # Всё ок — готовим объект
                workers_to_create.append(
                    Worker(
                        first_name=first_name,
                        middle_name=middle_name,
                        last_name=last_name,
                        email=email,
                        position=position,
                        created_by=request.user
                    )
                )

            # Если есть ошибки — не сохраняем ничего
            if errors:
                return Response({
                    "success_count": 0,
                    "errors": errors
                }, status=status.HTTP_400_BAD_REQUEST)

            # Сохраняем атомарно
            with transaction.atomic():
                Worker.objects.bulk_create(workers_to_create)

            logger.info(f"Импорт успешен: {len(workers_to_create)} работников добавлено пользователем {request.user.username}")

            return Response({
                "success_count": len(workers_to_create),
                "errors": []
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Ошибка при импорте: {str(e)}")
            return Response(
                {"error": "Некорректный файл Excel"},
                status=status.HTTP_400_BAD_REQUEST
            )
